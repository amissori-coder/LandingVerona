#!/usr/bin/env python3
"""Selezione delle aziende da invitare al convegno di Napoli del 2 ottobre.

Legge l'anagrafica Campania (xlsx) e produce un file di lavoro con le imprese
più coerenti con i temi dell'evento: rigenerazione di Bagnoli e America's Cup
2027, con Invitalia, UniCredit e Intesa Sanpaolo al tavolo.

Uso:
    python3 seleziona_aziende.py Campania.xlsx [-o lista-aziende-2-ottobre.xlsx]

I dati di input e di output restano fuori dal repository: contengono email e
telefoni delle aziende (vedi .gitignore).
"""

import argparse
import json
import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 1. Settori target
#
# I settori sono quelli indicati nella mail di Sergio: costruzioni e
# infrastrutture, impiantistica, bonifiche ambientali, nautica, portualità,
# turismo e ospitalità, ristorazione, trasporti e logistica, eventi,
# tecnologie digitali, sicurezza, energia e servizi.
#
# La fascia (A/B/C) misura quanto il settore è vicino al cuore dei due
# progetti: A = filiera diretta di Bagnoli e della Coppa America, B = settori
# che l'evento tocca in modo rilevante, C = filiere di fornitura e servizi.
# --------------------------------------------------------------------------

COSTRUZIONI = "Costruzioni e infrastrutture"
IMPIANTI = "Impiantistica"
AMBIENTE = "Bonifiche e ambiente"
NAUTICA = "Nautica e portualità"
TRASPORTI = "Trasporti e logistica"
TURISMO = "Turismo e ospitalità"
RISTORAZIONE = "Ristorazione"
EVENTI = "Eventi e comunicazione"
DIGITALE = "Tecnologie digitali"
SICUREZZA = "Sicurezza"
ENERGIA = "Energia"
SERVIZI = "Servizi alle imprese"
FILIERA = "Filiera e forniture tecniche"
INDUSTRIA = "Industria e meccanica"

# prefisso ATECO -> (settore, fascia). Vince il prefisso più lungo.
REGOLE_ATECO = {
    # --- Costruzioni, infrastrutture, sviluppo immobiliare ---------------
    "41": (COSTRUZIONI, "A"),
    "42": (COSTRUZIONI, "A"),
    "43": (COSTRUZIONI, "A"),
    "43.13": (COSTRUZIONI, "A"),
    "43.99": (COSTRUZIONI, "A"),
    "08.12": (COSTRUZIONI, "A"),
    "23.6": (COSTRUZIONI, "A"),
    "23.61": (COSTRUZIONI, "A"),
    "23.63": (COSTRUZIONI, "A"),
    "23.64": (COSTRUZIONI, "A"),
    "23.70": (COSTRUZIONI, "A"),
    "25.11": (COSTRUZIONI, "A"),
    "71.12": (COSTRUZIONI, "A"),
    "68": (COSTRUZIONI, "B"),
    "68.1": (COSTRUZIONI, "B"),
    "68.20": (COSTRUZIONI, "B"),
    # --- Impiantistica e manutenzione industriale ------------------------
    "43.21": (IMPIANTI, "A"),
    "43.22": (IMPIANTI, "A"),
    "43.29": (IMPIANTI, "A"),
    "33.12": (IMPIANTI, "A"),
    "33.20": (IMPIANTI, "A"),
    "28.22": (IMPIANTI, "A"),
    # --- Bonifiche, rifiuti, ciclo idrico --------------------------------
    "36": (AMBIENTE, "A"),
    "37": (AMBIENTE, "A"),
    "38": (AMBIENTE, "A"),
    "39": (AMBIENTE, "A"),
    "28.29.91": (AMBIENTE, "A"),
    "46.77": (AMBIENTE, "B"),
    "81.22": (AMBIENTE, "B"),
    # --- Nautica e portualità -------------------------------------------
    "30.11": (NAUTICA, "A"),
    "30.12": (NAUTICA, "A"),
    "33.15": (NAUTICA, "A"),
    "50": (NAUTICA, "A"),
    "52.22": (NAUTICA, "A"),
    "52.24.2": (NAUTICA, "A"),
    "47.64.2": (NAUTICA, "B"),
    # --- Trasporti e logistica -------------------------------------------
    "49": (TRASPORTI, "B"),
    "52": (TRASPORTI, "B"),
    "30.20": (TRASPORTI, "B"),
    "33.16": (TRASPORTI, "C"),
    "33.17": (TRASPORTI, "B"),
    "77.11": (TRASPORTI, "C"),
    # --- Turismo e ospitalità -------------------------------------------
    "55": (TURISMO, "B"),
    "79": (TURISMO, "B"),
    "91.02": (TURISMO, "C"),
    "91.03": (TURISMO, "C"),
    "91.04": (TURISMO, "C"),
    "93.29.2": (TURISMO, "B"),
    "96.04.2": (TURISMO, "C"),
    # --- Ristorazione -----------------------------------------------------
    "56": (RISTORAZIONE, "B"),
    # --- Eventi, sport, comunicazione ------------------------------------
    "82.3": (EVENTI, "A"),
    "90": (EVENTI, "B"),
    "90.03": (COSTRUZIONI, "B"),
    "93.11": (EVENTI, "C"),
    "93.19": (EVENTI, "A"),
    "93.29.9": (EVENTI, "B"),
    "77.39.94": (EVENTI, "B"),
    "73.1": (EVENTI, "B"),
    "73.11": (EVENTI, "B"),
    "73.2": (EVENTI, "C"),
    "70.21": (EVENTI, "C"),
    "58": (EVENTI, "C"),
    "59": (EVENTI, "C"),
    # --- Tecnologie digitali ----------------------------------------------
    "61": (DIGITALE, "B"),
    "62": (DIGITALE, "B"),
    "63": (DIGITALE, "B"),
    "72.1": (DIGITALE, "B"),
    "26.12": (DIGITALE, "C"),
    "26.30": (DIGITALE, "C"),
    "46.51": (DIGITALE, "C"),
    "47.41": (DIGITALE, "C"),
    "47.42": (DIGITALE, "C"),
    # --- Sicurezza ---------------------------------------------------------
    "80": (SICUREZZA, "B"),
    "74.90.21": (SICUREZZA, "B"),
    "71.20": (SICUREZZA, "B"),
    "32.99.11": (SICUREZZA, "C"),
    # --- Energia -----------------------------------------------------------
    "35": (ENERGIA, "A"),
    "19.20.3": (ENERGIA, "B"),
    "46.71": (ENERGIA, "B"),
    "47.3": (ENERGIA, "C"),
    "47.78.4": (ENERGIA, "C"),
    "27.32": (ENERGIA, "C"),
    "27.33": (ENERGIA, "C"),
    "27.90": (ENERGIA, "C"),
    # --- Servizi alle imprese ---------------------------------------------
    "81.1": (SERVIZI, "B"),
    "81.2": (SERVIZI, "B"),
    "81.21": (SERVIZI, "B"),
    "81.29": (SERVIZI, "C"),
    "81.3": (SERVIZI, "C"),
    "78": (SERVIZI, "C"),
    "82.2": (SERVIZI, "C"),
    "82.9": (SERVIZI, "C"),
    "82.99": (SERVIZI, "C"),
    "70.1": (SERVIZI, "C"),
    "70.22": (SERVIZI, "C"),
    "74.1": (SERVIZI, "C"),
    "74.10": (SERVIZI, "C"),
    # --- Filiera e forniture tecniche -------------------------------------
    "46.47.3": (FILIERA, "C"),
    "46.63": (FILIERA, "B"),
    "46.69.19": (FILIERA, "C"),
    "46.69.99": (FILIERA, "C"),
    "46.72": (FILIERA, "B"),
    "46.73": (FILIERA, "B"),
    "46.74": (FILIERA, "B"),
    "47.52.1": (FILIERA, "C"),
    "47.52.2": (FILIERA, "C"),
    "47.52.3": (FILIERA, "C"),
    "47.52.4": (FILIERA, "C"),
    "25.12": (FILIERA, "C"),
    "16.23": (FILIERA, "C"),
    "77.32": (FILIERA, "B"),
    "77.39.99": (FILIERA, "C"),
    # --- Industria e meccanica (potenziali fornitori di filiera) ----------
    "24": (INDUSTRIA, "C"),
    "25": (INDUSTRIA, "C"),
    "26": (INDUSTRIA, "C"),
    "27": (INDUSTRIA, "C"),
    "28": (INDUSTRIA, "C"),
    "29": (INDUSTRIA, "C"),
    "30.30": (INDUSTRIA, "C"),
    "32.50": (INDUSTRIA, "C"),
    "72.11": (INDUSTRIA, "C"),
}

FUORI_TARGET = "Fuori target"

# Aggancio fra settore e opportunità da citare nella mail di invito.
FOCUS = {
    COSTRUZIONI: "Bagnoli: opere civili, waterfront, rigenerazione urbana",
    IMPIANTI: "Bagnoli e siti di gara: impianti, reti, manutenzione",
    AMBIENTE: "Bagnoli: bonifica suoli e falda, gestione materiali",
    NAUTICA: "America's Cup 2027: cantieri, refit, marina e basi team",
    TRASPORTI: "Mobilità e logistica di cantiere e di evento",
    TURISMO: "America's Cup 2027: ricettività e flussi internazionali",
    RISTORAZIONE: "Ospitalità e ristorazione per l'evento e l'indotto",
    EVENTI: "Organizzazione, allestimenti e comunicazione dell'evento",
    DIGITALE: "Digitalizzazione di cantiere, evento e servizi al pubblico",
    SICUREZZA: "Sicurezza dei siti, vigilanza, collaudi e certificazioni",
    ENERGIA: "Reti, efficienza e forniture energetiche dei nuovi poli",
    SERVIZI: "Servizi integrati alle imprese della filiera",
    FILIERA: "Forniture tecniche per cantieri e infrastrutture",
    INDUSTRIA: "Subfornitura industriale per le filiere Bagnoli/Coppa America",
}

PESO_FASCIA = {"A": 40, "B": 32, "C": 22}

# Soglie di punteggio e quote minime per settore delle due fasce di invito.
SOGLIA_P1, QUOTA_P1 = 78, 10
SOGLIA_P2, QUOTA_P2 = 66, 25

# Comuni dell'area flegrea e della città di Napoli: il perimetro fisico dei
# due progetti.
COMUNI_AREA_PROGETTO = {
    "NAPOLI", "POZZUOLI", "BACOLI", "MONTE DI PROCIDA", "QUARTO", "GIUGLIANO IN CAMPANIA",
}


def norm_ateco(valore) -> str:
    if valore is None or (isinstance(valore, float) and math.isnan(valore)):
        return ""
    return str(valore).strip()


def classifica(ateco_cod: str):
    """Restituisce (settore, fascia) applicando il prefisso ATECO più lungo."""
    codice = norm_ateco(ateco_cod)
    if not codice:
        return FUORI_TARGET, None
    parti = codice.split(".")
    for taglio in range(len(parti), 0, -1):
        prefisso = ".".join(parti[:taglio])
        if prefisso in REGOLE_ATECO:
            return REGOLE_ATECO[prefisso]
    return FUORI_TARGET, None


def serie_storica(storico_json, etichetta: str):
    """Estrae {anno: valore} dal campo storico (lista di dizionari in JSON)."""
    if not isinstance(storico_json, str) or not storico_json.strip():
        return {}
    try:
        voci = json.loads(storico_json)
    except json.JSONDecodeError:
        return {}
    serie = {}
    for voce in voci if isinstance(voci, list) else []:
        if not isinstance(voce, dict):
            continue
        for chiave, valore in voce.items():
            anno = re.search(r"(19|20)\d{2}", str(chiave))
            if anno and etichetta.lower() in str(chiave).lower() and valore is not None:
                try:
                    serie[int(anno.group(0))] = float(valore)
                except (TypeError, ValueError):
                    pass
    return serie


def cagr_fatturato(riga):
    """Crescita media annua del fatturato sui dati disponibili."""
    serie = serie_storica(riga.get("storico_fatturato"), "Fatturato")
    ultimo = riga.get("ultimo_fatturato")
    anno_ultimo = None
    data_bilancio = riga.get("dataultimobilancio")
    if isinstance(data_bilancio, str):
        anno = re.search(r"(19|20)\d{2}", data_bilancio)
        if anno:
            anno_ultimo = int(anno.group(0))
    elif hasattr(data_bilancio, "year"):
        anno_ultimo = data_bilancio.year
    if ultimo is not None and not pd.isna(ultimo) and anno_ultimo:
        serie[anno_ultimo] = float(ultimo)
    if len(serie) < 2:
        return None
    anni = sorted(serie)
    primo, ultimo_anno = anni[0], anni[-1]
    v0, v1 = serie[primo], serie[ultimo_anno]
    if v0 <= 0 or v1 <= 0 or ultimo_anno == primo:
        return None
    return (v1 / v0) ** (1 / (ultimo_anno - primo)) - 1


def punteggio_dimensione(fatturato, dipendenti):
    """0-25: 'aziende qualificate e di buona dimensione'."""
    def fascia(valore, soglie):
        if valore is None or pd.isna(valore):
            return None
        for soglia, punti in soglie:
            if valore >= soglia:
                return punti
        return 3

    p_fat = fascia(fatturato, [(50e6, 25), (20e6, 22), (10e6, 19), (5e6, 15), (2e6, 10), (1e6, 6)])
    p_dip = fascia(dipendenti, [(250, 25), (100, 22), (50, 18), (25, 14), (10, 9), (5, 6)])
    if p_fat is None and p_dip is None:
        return 8.0
    if p_fat is None:
        return float(p_dip)
    if p_dip is None:
        return float(p_fat)
    return 0.6 * p_fat + 0.4 * p_dip


def punteggio_crescita(cagr):
    """0-15: chi cresce è chi investe e cerca finanza."""
    if cagr is None:
        return 6.0
    for soglia, punti in [(0.20, 15), (0.10, 13), (0.05, 11), (0.02, 9), (0.0, 7), (-0.05, 4), (-0.15, 2)]:
        if cagr >= soglia:
            return float(punti)
    return 0.0


def punteggio_solidita(utile, fatturato):
    """0-10: imprese 'positivamente riferite', con conti in ordine."""
    if utile is None or pd.isna(utile):
        return 3.0
    if utile <= 0:
        return 1.0
    punti = 6.0
    if fatturato and not pd.isna(fatturato) and fatturato > 0:
        margine = utile / fatturato
        if margine >= 0.05:
            punti += 4
        elif margine >= 0.02:
            punti += 2
    return punti


def punteggio_territorio(citta, provincia):
    """0-10: vicinanza al perimetro dei progetti."""
    citta = str(citta).strip().upper()
    provincia = str(provincia).strip()
    if citta in COMUNI_AREA_PROGETTO:
        return 10.0
    if provincia == "Napoli":
        return 8.0
    if provincia in ("Caserta", "Salerno"):
        return 5.0
    return 3.0


def descrizione_breve(ateco: str) -> str:
    testo = str(ateco or "").strip()
    if ":" in testo:
        testo = testo.split(":", 1)[1].strip()
    return testo[:120]


def euro(valore):
    if valore is None or pd.isna(valore):
        return "n.d."
    return f"{valore/1e6:.1f} M€" if valore >= 1e6 else f"{valore/1e3:.0f} k€"


def motivo(riga):
    pezzi = [riga["settore"], f"fatturato {euro(riga['ultimo_fatturato'])}"]
    if riga["dipendenti"] and not pd.isna(riga["dipendenti"]):
        pezzi.append(f"{int(riga['dipendenti'])} addetti")
    if riga["cagr"] is not None:
        pezzi.append(f"trend {riga['cagr']*100:+.0f}%/anno")
    pezzi.append(str(riga["citta"]).title())
    return " · ".join(pezzi)


def prepara(percorso_input: Path) -> pd.DataFrame:
    df = pd.read_excel(percorso_input)
    settori = df["ateco_cod"].map(lambda c: classifica(c))
    df["settore"] = [s for s, _ in settori]
    df["fascia"] = [f for _, f in settori]
    df["cagr"] = df.apply(cagr_fatturato, axis=1)

    df["p_settore"] = df["fascia"].map(lambda f: PESO_FASCIA.get(f, 0))
    df["p_dimensione"] = df.apply(lambda r: punteggio_dimensione(r["ultimo_fatturato"], r["dipendenti"]), axis=1)
    df["p_crescita"] = df["cagr"].map(punteggio_crescita)
    df["p_solidita"] = df.apply(lambda r: punteggio_solidita(r["ultimo_utile"], r["ultimo_fatturato"]), axis=1)
    df["p_territorio"] = df.apply(lambda r: punteggio_territorio(r["citta"], r["provincia"]), axis=1)
    df["punteggio"] = (
        df["p_settore"] + df["p_dimensione"] + df["p_crescita"] + df["p_solidita"] + df["p_territorio"]
    ).round(1)

    in_target = df[df["settore"] != FUORI_TARGET].copy()
    in_target = in_target.sort_values(["punteggio", "ultimo_fatturato"], ascending=[False, False])

    # Priorità: soglia di punteggio più una quota per settore, così nessuno
    # dei settori citati nella mail resta senza rappresentanza in sala.
    posizione = in_target.groupby("settore")["punteggio"].rank(method="first", ascending=False)
    p1 = (in_target["punteggio"] >= SOGLIA_P1) | (posizione <= QUOTA_P1)
    p2 = ~p1 & ((in_target["punteggio"] >= SOGLIA_P2) | (posizione <= QUOTA_P2))
    in_target["priorita"] = "3 - Riserva"
    in_target.loc[p2, "priorita"] = "2 - Estensione"
    in_target.loc[p1, "priorita"] = "1 - Priorità alta"
    in_target["posizione_settore"] = posizione.astype(int)
    in_target["focus"] = in_target["settore"].map(FOCUS)
    in_target["citta"] = in_target["citta"].astype(str).str.title()
    in_target["motivo"] = in_target.apply(motivo, axis=1)
    in_target = in_target.reset_index(drop=True)
    in_target.insert(0, "rank", range(1, len(in_target) + 1))
    return df, in_target


# --------------------------------------------------------------------------
# 2. Scrittura del file di lavoro
# --------------------------------------------------------------------------

FONT = "Arial"
BLU = "1F3864"
GRIGIO = "F2F2F2"

COLONNE = [
    ("rank", "#", 6),
    ("punteggio", "Punteggio", 11),
    ("priorita", "Priorità", 17),
    ("ragione_sociale", "Ragione sociale", 46),
    ("settore", "Settore target", 26),
    ("focus", "Opportunità da citare nell'invito", 46),
    ("attivita", "Attività (ATECO)", 50),
    ("ateco_cod", "ATECO", 10),
    ("citta", "Città", 22),
    ("provincia", "Provincia", 12),
    ("dipendenti", "Addetti", 9),
    ("ultimo_fatturato", "Fatturato (€)", 16),
    ("ultimo_utile", "Utile (€)", 14),
    ("cagr", "Trend fatturato", 15),
    ("anno_bilancio", "Anno bilancio", 13),
    ("email_valida", "Email", 34),
    ("telefono", "Telefono", 18),
    ("url", "Sito", 34),
    ("piva", "P.IVA", 14),
    ("motivo", "Sintesi", 60),
]


def intesta(ws, riga=1):
    for col, (_, etichetta, larghezza) in enumerate(COLONNE, start=1):
        cella = ws.cell(row=riga, column=col, value=etichetta)
        cella.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cella.fill = PatternFill("solid", fgColor=BLU)
        cella.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = larghezza
    ws.row_dimensions[riga].height = 30
    ws.freeze_panes = ws.cell(row=riga + 1, column=1)


def scrivi_righe(ws, dati: pd.DataFrame, riga_inizio=2):
    bordo = Side(style="thin", color="D9D9D9")
    for i, (_, r) in enumerate(dati.iterrows()):
        riga = riga_inizio + i
        for col, (chiave, _, _) in enumerate(COLONNE, start=1):
            valore = r.get(chiave)
            if isinstance(valore, float) and pd.isna(valore):
                valore = None
            cella = ws.cell(row=riga, column=col, value=valore)
            cella.font = Font(name=FONT, size=10)
            cella.border = Border(bottom=bordo)
            cella.alignment = Alignment(
                vertical="top", wrap_text=chiave in ("focus", "attivita", "ragione_sociale", "motivo")
            )
            if chiave in ("ultimo_fatturato", "ultimo_utile"):
                cella.number_format = "#,##0"
            elif chiave == "cagr":
                cella.number_format = "0.0%"
            elif chiave == "punteggio":
                cella.number_format = "0.0"
        if i % 2:
            for col in range(1, len(COLONNE) + 1):
                ws.cell(row=riga, column=col).fill = PatternFill("solid", fgColor=GRIGIO)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLONNE))}{riga_inizio + len(dati) - 1}"


def foglio_lista(wb, titolo: str, dati: pd.DataFrame):
    ws = wb.create_sheet(titolo)
    intesta(ws)
    if len(dati):
        scrivi_righe(ws, dati)
    return ws


def foglio_metodologia(wb, totale, in_target, conteggi):
    ws = wb.create_sheet("Metodologia", 0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96
    righe = [
        ("Convegno del 2 ottobre 2026 - Napoli", ""),
        ("Oggetto", "Selezione delle aziende campane da invitare, a partire dall'anagrafica Campania"),
        ("Criteri", "Mail di Sergio: settori collegati a Bagnoli e America's Cup 2027, imprese che investono, "
                    "cercano finanza o vogliono entrare nelle filiere, di buona dimensione"),
        ("", ""),
        ("Come è costruito il punteggio (0-100)", ""),
        ("Settore (max 40)", "Fascia A = filiera diretta dei due progetti (costruzioni, impiantistica, bonifiche, "
                             "nautica e portualità, energia, organizzazione eventi): 40 punti. "
                             "Fascia B = settori toccati in modo rilevante: 32 punti. Fascia C = filiere di "
                             "fornitura e servizi: 22 punti"),
        ("Dimensione (max 25)", "Ultimo fatturato disponibile (peso 60%) e numero di addetti (peso 40%)"),
        ("Crescita (max 15)", "Crescita media annua del fatturato sullo storico disponibile: chi cresce è chi "
                              "investe e ha bisogno di finanza. In assenza di storico si assegna un valore neutro"),
        ("Solidità (max 10)", "Utile dell'ultimo bilancio e margine sul fatturato"),
        ("Territorio (max 10)", "Napoli e area flegrea 10, resto provincia di Napoli 8, Caserta e Salerno 5, "
                                "Avellino e Benevento 3"),
        ("", ""),
        ("Fasce di priorità", ""),
        ("1 - Priorità alta", f"Punteggio >= {SOGLIA_P1} oppure fra le prime {QUOTA_P1} del proprio settore. "
                               "Prima ondata di mail e telefonate di follow-up"),
        ("2 - Estensione", f"Punteggio >= {SOGLIA_P2} oppure fra le prime {QUOTA_P2} del proprio settore. "
                           "Seconda ondata di mail"),
        ("3 - Riserva", "Le altre aziende in target, da usare per completare la sala"),
        ("Quota per settore", "La quota garantisce che tutti i settori indicati nella mail siano rappresentati: "
                              "alberghi, ristoranti e società di eventi hanno fatturati più bassi delle imprese "
                              "di costruzioni e senza quota resterebbero fuori dalla prima ondata"),
        ("", ""),
        ("Numeri", ""),
        ("Aziende nel file di partenza", totale),
        ("Aziende nei settori target", in_target),
        ("Priorità 1", conteggi.get("1 - Priorità alta", 0)),
        ("Priorità 2", conteggi.get("2 - Estensione", 0)),
        ("Priorità 3", conteggi.get("3 - Riserva", 0)),
        ("", ""),
        ("Limiti dei dati", "L'anagrafica non riporta né i piani di investimento né le referenze: il punteggio "
                            "usa dimensione, crescita e solidità come indizi. I bilanci più recenti sono al "
                            "2021, quindi il trend fotografa gli anni disponibili. La verifica di merito "
                            "('aziende positivamente riferite') resta al gruppo"),
        ("Esclusioni", "Restano fuori agroalimentare, sanità, istruzione, moda, commercio di autoveicoli, "
                       "farmacie e commercio al dettaglio generico: settori non citati nella mail"),
    ]
    for i, (etichetta, valore) in enumerate(righe, start=1):
        a = ws.cell(row=i, column=1, value=etichetta)
        b = ws.cell(row=i, column=2, value=valore)
        a.font = Font(name=FONT, size=10, bold=True, color=BLU if etichetta and not valore else "000000")
        b.font = Font(name=FONT, size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(name=FONT, size=13, bold=True, color=BLU)
    return ws


def foglio_riepilogo(wb, in_target: pd.DataFrame):
    """Conteggi per settore. Valori gia' calcolati: il file nasce come lista di
    lavoro, non come modello da ricalcolare."""
    ws = wb.create_sheet("Riepilogo per settore")
    intestazioni = ["Settore target", "Aziende in target", "Priorità 1", "Priorità 2",
                    "Priorità 3", "Fatturato complessivo (€)", "Addetti complessivi"]
    for col, testo in enumerate(intestazioni, start=1):
        cella = ws.cell(row=1, column=col, value=testo)
        cella.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cella.fill = PatternFill("solid", fgColor=BLU)
        cella.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 32 if col == 1 else 18
    ws.row_dimensions[1].height = 30

    riepilogo = (
        in_target.assign(uno=1)
        .pivot_table(
            index="settore",
            columns="priorita",
            values="uno",
            aggfunc="sum",
            fill_value=0,
        )
        .reindex(columns=["1 - Priorità alta", "2 - Estensione", "3 - Riserva"], fill_value=0)
    )
    aggregati = in_target.groupby("settore").agg(
        aziende=("piva", "size"),
        fatturato=("ultimo_fatturato", "sum"),
        addetti=("dipendenti", "sum"),
    )
    tabella = aggregati.join(riepilogo).sort_values("aziende", ascending=False)

    for i, (settore, r) in enumerate(tabella.iterrows(), start=2):
        valori = [settore, int(r["aziende"]), int(r["1 - Priorità alta"]),
                  int(r["2 - Estensione"]), int(r["3 - Riserva"]),
                  float(r["fatturato"]), int(r["addetti"])]
        for col, valore in enumerate(valori, start=1):
            cella = ws.cell(row=i, column=col, value=valore)
            cella.font = Font(name=FONT, size=10)
            if col >= 6:
                cella.number_format = "#,##0"
    riga_tot = len(tabella) + 2
    totali = ["Totale", int(tabella["aziende"].sum()), int(tabella["1 - Priorità alta"].sum()),
              int(tabella["2 - Estensione"].sum()), int(tabella["3 - Riserva"].sum()),
              float(tabella["fatturato"].sum()), int(tabella["addetti"].sum())]
    for col, valore in enumerate(totali, start=1):
        cella = ws.cell(row=riga_tot, column=col, value=valore)
        cella.font = Font(name=FONT, size=10, bold=True)
        if col >= 6:
            cella.number_format = "#,##0"
    nota = ws.cell(row=riga_tot + 2, column=1,
                   value="Conteggi calcolati sul foglio «Lista completa in target». "
                         "Fatturato e addetti sono somme degli ultimi dati di bilancio disponibili.")
    nota.font = Font(name=FONT, size=9, italic=True)
    return ws


def costruisci(df_completo: pd.DataFrame, in_target: pd.DataFrame, percorso_output: Path):
    wb = Workbook()
    wb.remove(wb.active)

    p1 = in_target[in_target["priorita"] == "1 - Priorità alta"]
    p2 = in_target[in_target["priorita"] == "2 - Estensione"]
    p3 = in_target[in_target["priorita"] == "3 - Riserva"]

    nome_lista = "Lista completa in target"
    foglio_lista(wb, "1 - Priorità alta", p1)
    foglio_lista(wb, "2 - Estensione", p2)
    foglio_lista(wb, "3 - Riserva", p3)
    foglio_lista(wb, nome_lista, in_target)
    foglio_riepilogo(wb, in_target)
    foglio_metodologia(wb, len(df_completo), len(in_target), in_target["priorita"].value_counts().to_dict())

    wb.save(percorso_output)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="anagrafica xlsx di partenza")
    parser.add_argument("-o", "--output", type=Path, default=Path("lista-aziende-2-ottobre.xlsx"))
    args = parser.parse_args()

    df_completo, in_target = prepara(args.input)
    in_target["attivita"] = in_target["ateco"].map(descrizione_breve)
    in_target["anno_bilancio"] = in_target["dataultimobilancio"].map(
        lambda d: int(re.search(r"(19|20)\d{2}", str(d)).group(0)) if re.search(r"(19|20)\d{2}", str(d)) else None
    )
    costruisci(df_completo, in_target, args.output)

    conteggi = in_target["priorita"].value_counts()
    print(f"Aziende lette: {len(df_completo)}")
    print(f"Aziende nei settori target: {len(in_target)}")
    for etichetta in ("1 - Priorità alta", "2 - Estensione", "3 - Riserva"):
        print(f"  {etichetta}: {conteggi.get(etichetta, 0)}")
    print(f"File scritto: {args.output}")


if __name__ == "__main__":
    main()
